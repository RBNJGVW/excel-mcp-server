import logging
import os
from typing import Any, Dict, List, Optional

from mcp.server.fastmcp import FastMCP

from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.data import write_data

# Import exceptions
from excel_mcp.exceptions import (
    CalculationError,
    ChartError,
    DataError,
    FormattingError,
    PivotError,
    SheetError,
    ValidationError,
    WorkbookError,
)
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_cols,
    delete_rows,
    delete_sheet,
    get_merged_ranges,
    insert_cols,
    insert_row,
    merge_range,
    rename_sheet,
    unmerge_range,
)
from excel_mcp.storage_backend import get_storage
from excel_mcp.tables import create_excel_table as create_table_impl

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
)
from excel_mcp.validation import (
    validate_range_in_sheet_operation as validate_range_impl,
)
from excel_mcp.workbook import get_workbook_info

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Backend de almacenamiento
STORAGE = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files",
)


def _logical_name(filepath: str) -> str:
    """
    Devuelve un nombre lógico para el backend.
    - Si te pasan un path absoluto, usamos solo el basename (misma semántica que antes).
    - Normalizamos separadores y quitamos prefijos redundantes cuando el backend es Blob.
    """
    n = (filepath or "").strip()
    if os.path.isabs(n):
        n = os.path.basename(n)
    n = n.replace("\\", "/").lstrip("/")

    _ensure_storage()  # asegura STORAGE inicializado
    if hasattr(STORAGE, "normalize_name"):
        n = STORAGE.normalize_name(n)
    return n


def _ensure_storage():
    global STORAGE, EXCEL_FILES_PATH
    if STORAGE is None:
        STORAGE = get_storage(EXCEL_FILES_PATH or "./excel_files")


def _read_call(filepath: str, func, *args, **kwargs):
    """Ejecuta `func(local_path, ...)` leyendo desde FS local o Blob."""
    _ensure_storage()
    name = _logical_name(filepath)
    with STORAGE.local_read(name) as local_path:
        return func(local_path, *args, **kwargs)


def _write_call(filepath: str, func, *args, **kwargs):
    """Ejecuta `func(local_path, ...)` escribiendo y subiendo a Blob si procede."""
    _ensure_storage()
    name = _logical_name(filepath)
    with STORAGE.local_write(name) as local_path:
        return func(local_path, *args, **kwargs)


def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.

    Args:
        filename: Name of Excel file

    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(
            f"Invalid filename: {filename}, must be an absolute path when not in SSE mode"
        )

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(EXCEL_FILES_PATH, filename)


@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        # First validate the formula
        validation = _read_call(
            filepath, validate_formula_impl, sheet_name, cell, formula
        )
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"

        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl

        result = _write_call(filepath, apply_formula_impl, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise


@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:

        result = _read_call(filepath, validate_formula_impl, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise


@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
) -> str:
    """Apply formatting to a range of cells."""
    try:

        from excel_mcp.formatting import format_range as format_range_func

        # Convert None values to appropriate defaults for the underlying function
        _write_call(
            filepath,
            format_range_func,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,
            font_color=font_color,
            bg_color=bg_color,
            border_style=border_style,
            border_color=border_color,
            number_format=number_format,
            alignment=alignment,
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,
            conditional_format=conditional_format,
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise


@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False,
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only

    Returns:
    JSON string containing structured cell data with validation metadata.
    Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:

        from excel_mcp.data import read_excel_range_with_metadata

        result = _read_call(
            filepath, read_excel_range_with_metadata, sheet_name, start_cell, end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"

        # Return as formatted JSON string
        import json

        return json.dumps(result, indent=2, default=str)

    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise


@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"

    """
    try:

        result = _write_call(filepath, write_data, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise


@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:

        from excel_mcp.workbook import create_workbook as create_workbook_impl

        _write_call(filepath, create_workbook_impl)
        return f"Created workbook at {filepath}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise


@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:

        from excel_mcp.workbook import create_sheet as create_worksheet_impl

        result = _write_call(filepath, create_worksheet_impl, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise


@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> str:
    """Create chart in worksheet."""
    try:

        result = _write_call(
            filepath,
            create_chart_impl,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis,
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise


@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean",
) -> str:
    """Create pivot table in worksheet."""
    try:

        result = _write_call(
            filepath,
            create_pivot_table_impl,
            sheet_name,
            data_range,
            rows,
            values,
            columns or [],
            agg_func,
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise


@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9",
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:

        result = _write_call(
            filepath,
            create_table_impl,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style,
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise


@mcp.tool()
def copy_worksheet(filepath: str, source_sheet: str, target_sheet: str) -> str:
    """Copy worksheet within workbook."""
    try:

        result = _write_call(filepath, copy_sheet, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise


@mcp.tool()
def delete_worksheet(filepath: str, sheet_name: str) -> str:
    """Delete worksheet from workbook."""
    try:

        result = _write_call(filepath, delete_sheet, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise


@mcp.tool()
def rename_worksheet(filepath: str, old_name: str, new_name: str) -> str:
    """Rename worksheet in workbook."""
    try:

        result = _write_call(filepath, rename_sheet, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise


@mcp.tool()
def get_workbook_metadata(filepath: str, include_ranges: bool = False) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        result = _read_call(filepath, get_workbook_info, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise


@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:

        result = _write_call(filepath, merge_range, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise


@mcp.tool()
def unmerge_cells(
    filepath: str, sheet_name: str, start_cell: str, end_cell: str
) -> str:
    """Unmerge a range of cells."""
    try:

        result = _write_call(filepath, unmerge_range, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise


@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        return str(_read_call(filepath, get_merged_ranges, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise


@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None,
) -> str:
    """Copy a range of cells to another location."""
    try:

        from excel_mcp.sheet import copy_range_operation

        result = _write_call(
            filepath,
            copy_range_operation,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name,
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise


@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up",
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:

        from excel_mcp.sheet import delete_range_operation

        result = _write_call(
            filepath,
            delete_range_operation,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction,
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise


@mcp.tool()
def validate_excel_range(
    filepath: str, sheet_name: str, start_cell: str, end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = _read_call(filepath, validate_range_impl, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise


@mcp.tool()
def get_data_validation_info(filepath: str, sheet_name: str) -> str:
    """
    Get all data validation rules in a worksheet.

    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet

    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:

        def _fn(local_path: str, _sheet: str):
            from openpyxl import load_workbook

            from excel_mcp.cell_validation import get_all_validation_ranges

            wb = load_workbook(local_path, read_only=False)
            if _sheet not in wb.sheetnames:
                wb.close()
                return {"error": f"Sheet '{_sheet}' not found"}
            ws = wb[_sheet]
            validations = get_all_validation_ranges(ws)
            wb.close()
            return {"sheet_name": _sheet, "validation_rules": validations}

        result = _read_call(filepath, _fn, sheet_name)
        if "error" in result:
            return f"Error: {result['error']}"
        import json

        if not result["validation_rules"]:
            return "No data validation rules found in this worksheet"
        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise


@mcp.tool()
def insert_rows(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> str:
    """Insert one or more rows starting at the specified row."""
    try:

        result = _write_call(filepath, insert_row, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise


@mcp.tool()
def insert_columns(
    filepath: str, sheet_name: str, start_col: int, count: int = 1
) -> str:
    """Insert one or more columns starting at the specified column."""
    try:

        result = _write_call(filepath, insert_cols, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise


@mcp.tool()
def delete_sheet_rows(
    filepath: str, sheet_name: str, start_row: int, count: int = 1
) -> str:
    """Delete one or more rows starting at the specified row."""
    try:

        result = _write_call(filepath, delete_rows, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise


@mcp.tool()
def delete_sheet_columns(
    filepath: str, sheet_name: str, start_col: int, count: int = 1
) -> str:
    """Delete one or more columns starting at the specified column."""
    try:

        result = _write_call(filepath, delete_cols, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise


@mcp.tool()
def list_backend_files(pattern: Optional[str] = "*.xlsx") -> str:
    try:
        _ensure_storage()
        names = STORAGE.list_names(pattern)
        return "\n".join(names) if names else "(vacío)"
    except Exception as e:
        return f"Error: {e}"


def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH, STORAGE
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    STORAGE = get_storage(EXCEL_FILES_PATH)

    # Solo crear carpeta si es FS local
    if not STORAGE.is_blob:
        os.makedirs(EXCEL_FILES_PATH, exist_ok=True)

    try:
        logger.info(
            f"Starting Excel MCP server with SSE transport (files base: {EXCEL_FILES_PATH})"
        )
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")


def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH, STORAGE
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    STORAGE = get_storage(EXCEL_FILES_PATH)

    # Solo crear carpeta si es FS local
    if not STORAGE.is_blob:
        os.makedirs(EXCEL_FILES_PATH, exist_ok=True)

    try:
        logger.info(
            f"Starting Excel MCP server with streamable HTTP transport (files base: {EXCEL_FILES_PATH})"
        )
        mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")


def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode

    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
